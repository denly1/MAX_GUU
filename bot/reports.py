"""Генерация Excel-выгрузок (openpyxl)."""

from __future__ import annotations

from datetime import datetime
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Font

from . import config, repo
from .database import get_conn


def _new_path(prefix: str) -> Path:
    config.ensure_dirs()
    ts = datetime.now().strftime("%Y%m%d_%H%M%S")
    return config.EXPORTS_DIR / f"{prefix}_{ts}.xlsx"


def _autoheader(ws, headers: list[str]) -> None:
    ws.append(headers)
    for cell in ws[1]:
        cell.font = Font(bold=True)


def export_task_choices() -> Path:
    """Выгрузка по выбору проектов (раздел 2.5)."""
    headers = [
        "Фамилия", "Имя", "Отчество", "Институт", "Курс",
        "Образовательная программа", "Группа", "Роль в команде", "Команда",
        "Наименование выбранного проекта", "Наименование социального партнера",
    ]
    wb = Workbook()
    ws = wb.active
    ws.title = "Выбор проектов"
    _autoheader(ws, headers)

    rows = get_conn().execute(
        """
        SELECT u.last_name, u.first_name, u.patronymic, u.institute, u.course,
               u.education_program, u.group_name, tm.role_in_team,
               t.name AS team_name, t.task_id
        FROM team_members tm
        JOIN users u ON u.user_id = tm.user_id
        JOIN teams t ON t.id = tm.team_id
        ORDER BY t.name, tm.role_in_team
        """
    ).fetchall()

    role_label = {"leader": "Лидер команды", "member": "Участник команды"}
    for r in rows:
        task = repo.get_task(r["task_id"]) if r["task_id"] else None
        ws.append([
            r["last_name"], r["first_name"], r["patronymic"], r["institute"],
            r["course"], r["education_program"], r["group_name"],
            role_label.get(r["role_in_team"], r["role_in_team"]),
            r["team_name"],
            task["title"] if task else "",
            task["partner_name"] if task else "",
        ])

    path = _new_path("vybor_proektov")
    wb.save(path)
    return path


def export_tasks() -> Path:
    """Выгрузка списка задач (раздел 2.11)."""
    headers = [
        "Номер", "Краткое название проекта", "Наименование социального партнера",
        "Краткое описание", "Количество команд",
    ]
    wb = Workbook()
    ws = wb.active
    ws.title = "Задачи"
    _autoheader(ws, headers)

    for t in repo.list_tasks():
        ws.append([
            t["number"], t["title"], t["partner_name"],
            t["description"], t["max_teams"],
        ])

    path = _new_path("zadachi")
    wb.save(path)
    return path
